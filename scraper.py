"""
scraper.py

Scrapes SPMCIL tenders and stores them in an Excel file.
"""

from datetime import datetime
import os
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook, load_workbook
from urllib.parse import urljoin
from config import TENDER_SITES, logger

EXCEL_FILE = "spmcil_tenders.xlsx"


class TenderScraper:

    def __init__(self):

        self.session = requests.Session()

        retries = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"],
        )

        adapter = HTTPAdapter(max_retries=retries)

        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)

        self.headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/137.0 Safari/537.36"
            )
        }

    # --------------------------------------------------------

    def fetch_page(self, url ):

        logger.info(f"Fetching {url}")

        response = self.session.get(
            url,
            headers=self.headers,
            timeout=30
        )

        response.raise_for_status()
        
        with open("debug.html", "w", encoding="utf-8") as f:
            f.write(response.text)

        return response.text

    # --------------------------------------------------------

    from urllib.parse import urljoin


    def  parse_corporate_table(self, html):

        soup = BeautifulSoup(html, "lxml")

        table = soup.find("table")

        if table is None:
            raise Exception("Tender table not found.")

        rows = table.find_all("tr")

        tenders = []

        base_url = "https://www.spmcil.com"

        for row in rows[1:]:

            cols = row.find_all("td")

            if len(cols) < 8:
                continue

            # -----------------------------
            # Corrigendum
            # -----------------------------

            corrigendum_text = ""
            corrigendum_url = ""

            corrigendum_link = cols[6].find("a")

            if corrigendum_link:

                corrigendum_text = corrigendum_link.get_text(
                    " ",
                    strip=True
                )

                corrigendum_url = urljoin(
                    base_url,
                    corrigendum_link.get("href", "")
                )

            # -----------------------------
            # Tender Document
            # -----------------------------

            tender_doc_text = ""
            tender_doc_url = ""

            tender_doc_link = cols[7].find("a")

            if tender_doc_link:

                tender_doc_text = tender_doc_link.get_text(
                    " ",
                    strip=True
                )

                tender_doc_url = urljoin(
                    base_url,
                    tender_doc_link.get("href", "")
                )

            tender = {

                "Unit Name": cols[1].get_text(" ", strip=True),

                "Tender Number": cols[2].get_text(strip=True),

                "Tender Title": cols[3].get_text(
                    " ",
                    strip=True
                ),

                "Publishing Date": cols[4].get_text(
                    " ",
                    strip=True
                ),

                "Closing Date": cols[5].get_text(
                    " ",
                    strip=True
                ),

                "Corrigendum": corrigendum_text,

                "Corrigendum URL": corrigendum_url,

                "Tender Document": tender_doc_text,

                "Tender Document URL": tender_doc_url,

                "Scraped At": datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                ),
            }

            tenders.append(tender)

        logger.info(f"Fetched {len(tenders)} tenders")

        return tenders


    def parse_unit_table(self, html, site):

        soup = BeautifulSoup(html, "lxml")

        table = soup.find("table")

        if table is None:
            raise Exception("Tender table not found.")

        rows = table.find_all("tr")

        tenders = []

        for row in rows[1:]:

            cols = row.find_all("td")

            # Skip invalid rows
            if len(cols) < 7:
                continue

            # ---------------------------------
            # Tender Document
            # ---------------------------------

            tender_doc = ""
            tender_doc_url = ""

            link = cols[0].find("a")

            if link:
                tender_doc = link.get_text(" ", strip=True)
                tender_doc_url = urljoin(
                    site["url"],
                    link.get("href", "")
                )

            # ---------------------------------
            # Corrigendum
            # ---------------------------------

            corr_text = ""
            corr_url = ""

            link = cols[6].find("a")

            if link:
                corr_text = link.get_text(" ", strip=True)
                corr_url = urljoin(
                    site["url"],
                    link.get("href", "")
                )

            tender = {

                "Unit Name": cols[1].get_text(" ", strip=True),

                "Tender Number": cols[2].get_text(" ", strip=True),

                "Tender Title": cols[3].get_text(" ", strip=True),

                "Publishing Date": cols[4].get_text(" ", strip=True),

                "Closing Date": cols[5].get_text(" ", strip=True),

                "Tender Document": tender_doc,

                "Tender Document URL": tender_doc_url,

                "Corrigendum": corr_text,

                "Corrigendum URL": corr_url,

                "Scraped At": datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            }

            tenders.append(tender)

        logger.info(
            f"{site['name']} -> Parsed {len(tenders)} tenders"
        )

        return tenders

    # --------------------------------------------------------

    def create_excel(self):

        if os.path.exists(EXCEL_FILE):
            return

        wb = Workbook()

        ws = wb.active

        ws.title = "Tenders"

        ws.append([
            "Source",
            "Unit Name",
            "Tender Number",
            "Tender Title",
            "Publishing Date",
            "Closing Date",
            "Tender Document",
            "Tender Document URL",
            "Corrigendum",
            "Corrigendum URL",
            "Scraped At"
        ])

        wb.save(EXCEL_FILE)

        logger.info("Excel created.")

    # --------------------------------------------------------

    def get_existing_tenders(self):

        wb = load_workbook(EXCEL_FILE)

        ws = wb.active

        existing = set()

        for row in ws.iter_rows(min_row=2, values_only=True):

            source = str(row[0]).strip() if row[0] else ""

            tender_no = str(row[2]).strip() if row[2] else ""

            if source and tender_no:

                existing.add(f"{source}|{tender_no}")

        wb.close()

        return existing

    # --------------------------------------------------------

    def save_to_excel(self, tenders):

        self.create_excel()

        existing = self.get_existing_tenders()

        wb = load_workbook(EXCEL_FILE)

        ws = wb.active

        new_count = 0

        for tender in tenders:

            key = f"{tender['Source']}|{tender['Tender Number']}"

            if key in existing:
                continue

            ws.append([
                tender["Source"],
                tender["Unit Name"],
                tender["Tender Number"],
                tender["Tender Title"],
                tender["Publishing Date"],
                tender["Closing Date"],
                tender["Tender Document"],
                tender["Tender Document URL"],
                tender["Corrigendum"],
                tender["Corrigendum URL"],
                tender["Scraped At"]
            ])
            existing.add(key) 
            
            new_count += 1

        wb.save(EXCEL_FILE)

        wb.close()

        logger.info(f"{new_count} new tenders inserted.")

        print(f"Inserted {new_count} new tenders into Excel.")

    # --------------------------------------------------------

    def scrape(self, site):

        html = self.fetch_page(site["url"])

        if site["type"] == "corporate":
            tenders = self.parse_corporate_table(html)
        else:
            tenders = self.parse_unit_table(html, site)

        # Add source information to every tender
        for tender in tenders:
            tender["Source"] = site["name"]
            tender["Source URL"] = site["url"]

        return tenders


# --------------------------------------------------------

def get_tenders():

    scraper = TenderScraper()

    return scraper.scrape()


# --------------------------------------------------------

if __name__ == "__main__":

    tenders = get_tenders()

    print(f"\nTotal Scraped : {len(tenders)}")

    for tender in tenders:
        print(tender)