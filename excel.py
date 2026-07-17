import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

from config import logger

EXCEL_FILE = "spmcil_tenders.xlsx"


class ExcelManager:

    def __init__(self):
        self.file = EXCEL_FILE

        if not os.path.exists(self.file):
            self.create_file()

    def create_file(self):

        wb = Workbook()
        ws = wb.active
        ws.title = "Tenders"

        ws.append([
            "Unit Name",
            "Tender Number",
            "Tender Title",
            "Publishing Date",
            "Closing Date",
            "Scraped At"
        ])

        wb.save(self.file)

        logger.info("Excel file created.")

    def get_existing_tenders(self):

        wb = load_workbook(self.file)
        ws = wb.active

        tenders = set()

        for row in ws.iter_rows(min_row=2, values_only=True):

            if row[1]:
                tenders.add(str(row[1]).strip())

        wb.close()

        return tenders

    def append_tender(self, tender):

        wb = load_workbook(self.file)
        ws = wb.active

        ws.append([
            tender["Unit Name"],
            tender["Tender Number"],
            tender["Tender Title"],
            tender["Publishing Date"],
            tender["Closing Date"],
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])

        wb.save(self.file)
        wb.close()

        logger.info(
            f'Inserted {tender["Tender Number"]}'
        )

    def append_many(self, tenders):

        if not tenders:
            return

        wb = load_workbook(self.file)
        ws = wb.active

        for tender in tenders:

            ws.append([
                tender["Unit Name"],
                tender["Tender Number"],
                tender["Tender Title"],
                tender["Publishing Date"],
                tender["Closing Date"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])

        wb.save(self.file)
        wb.close()

        logger.info(f"{len(tenders)} rows added.")